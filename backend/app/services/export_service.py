"""
ScholarVision — Service Export Excel (openpyxl)
ERRIHANI Faiza — ENSET Média 2025/2026
"""
from __future__ import annotations
import io
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY_HEX="0B1437"; EMERALD_HEX="10B981"; ROSE_HEX="F43F5E"
ORANGE_HEX="F97316"; BLUE_HEX="1A56DB"; GRAY_LIGHT="F8FAFC"; WHITE_HEX="FFFFFF"; GRAY_MID="E2E8F0"

def _hfill(h=NAVY_HEX): return PatternFill("solid", fgColor=h)
def _afill(even): return PatternFill("solid", fgColor=GRAY_LIGHT if even else WHITE_HEX)
def _hfont(): return Font(name="Calibri", bold=True, size=11, color="FFFFFF")
def _bfont(bold=False, color="000000"): return Font(name="Calibri", bold=bold, size=10, color=color)
def _border():
    t = Side(style="thin", color=GRAY_MID)
    return Border(left=t, right=t, top=t, bottom=t)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def _colw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def _header_row(ws, headers, widths, row=1):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _hfill(); cell.font = _hfont()
        cell.alignment = _center(); cell.border = _border()
        _colw(ws, col, w)
    ws.row_dimensions[row].height = 30

def _data_row(ws, row, n, even):
    for col in range(1, n+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _afill(even); cell.font = _bfont()
        cell.border = _border(); cell.alignment = _left()
    ws.row_dimensions[row].height = 22

def _title_row(ws, title, n):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = PatternFill("solid", fgColor=NAVY_HEX)
    cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    cell.alignment = _center(); ws.row_dimensions[1].height = 36

def _freeze_filter(ws, hrow, n):
    ws.freeze_panes = ws.cell(row=hrow+1, column=1)
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(n)}{hrow}"

def _save(wb, output_path):
    if output_path:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path); return output_path
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


class ExcelExporter:

    def export_students_list(self, students, output_path=None):
        wb = Workbook(); ws = wb.active; ws.title = "Élèves"
        headers = ["N° Élève","Nom","Prénom","Classe","Niveau","Date naissance",
                   "Genre","Parent 1","Tél Parent","Email Parent","Ville","Statut","Date inscription"]
        widths = [14,18,18,12,10,16,8,22,16,26,14,10,16]
        _header_row(ws, headers, widths)
        for i, s in enumerate(students, 2):
            row = [s.get("student_number",""), s.get("last_name","").upper(), s.get("first_name",""),
                   s.get("class_name",""), s.get("level",""), s.get("date_of_birth",""),
                   s.get("gender",""), s.get("parent1_name",""), s.get("parent1_phone",""),
                   s.get("parent1_email",""), s.get("city",""),
                   "Actif" if s.get("is_active") else "Inactif", s.get("enrollment_date","")]
            for col, val in enumerate(row, 1): ws.cell(row=i, column=col, value=val)
            _data_row(ws, i, len(headers), i % 2 == 0)
            sc = ws.cell(row=i, column=12)
            sc.fill = PatternFill("solid", fgColor="ECFDF5" if s.get("is_active") else "FFF1F2")
            sc.font = Font(name="Calibri", bold=True, size=10, color=EMERALD_HEX if s.get("is_active") else ROSE_HEX)
            sc.alignment = _center()
        _title_row(ws, f"Liste des élèves — ScholarVision ({len(students)} élèves)", len(headers))
        _freeze_filter(ws, 2, len(headers))
        return _save(wb, output_path)

    def export_attendance_report(self, attendance_data, report_date=None, output_path=None):
        wb = Workbook(); ws = wb.active; ws.title = "Présences"
        headers = ["Classe","N° Élève","Nom","Prénom","Statut","Heure","Méthode","Score IA","Notes"]
        widths = [10,13,18,18,12,10,14,10,30]
        _header_row(ws, headers, widths)
        STATUS_C = {"present":("ECFDF5",EMERALD_HEX),"absent":("FFF1F2",ROSE_HEX),
                    "retard":("FFF7ED",ORANGE_HEX),"justified":("EFF6FF",BLUE_HEX)}
        STATUS_L = {"present":"✓ Présent","absent":"✗ Absent","retard":"⚡ Retard","justified":"✓ Justifié"}
        for i, att in enumerate(attendance_data, 2):
            status = att.get("status","")
            row = [att.get("class_name",""), att.get("student_number",""),
                   att.get("last_name","").upper(), att.get("first_name",""),
                   STATUS_L.get(status, status),
                   str(att.get("detected_at",""))[:5] if att.get("detected_at") else "",
                   att.get("detection_method","manuel"),
                   f"{att.get('confidence_score',0)*100:.1f}%" if att.get("confidence_score") else "—",
                   att.get("notes","")]
            for col, val in enumerate(row, 1): ws.cell(row=i, column=col, value=val)
            _data_row(ws, i, len(headers), i % 2 == 0)
            bg, fg = STATUS_C.get(status, (GRAY_LIGHT,"000000"))
            sc = ws.cell(row=i, column=5)
            sc.fill = PatternFill("solid", fgColor=bg)
            sc.font = Font(name="Calibri", bold=True, size=10, color=fg)
            sc.alignment = _center()
        date_str = str(report_date) if report_date else date.today().isoformat()
        _title_row(ws, f"Rapport de présences — {date_str}", len(headers))
        _freeze_filter(ws, 2, len(headers))
        return _save(wb, output_path)

    def export_financial_report(self, fees_data, month=None, output_path=None):
        wb = Workbook(); ws = wb.active; ws.title = "Finances"
        headers = ["N° Élève","Nom","Classe","Type frais","Total (MAD)","Payé (MAD)",
                   "Reste (MAD)","Statut","Échéance","Dernier paiement"]
        widths = [13,22,10,16,14,14,14,12,14,18]
        _header_row(ws, headers, widths)
        STATUS_C = {"paid":("ECFDF5",EMERALD_HEX),"partial":("FFF7ED",ORANGE_HEX),
                    "pending":("EFF6FF",BLUE_HEX),"overdue":("FFF1F2",ROSE_HEX)}
        STATUS_L = {"paid":"Payé","partial":"Partiel","pending":"En attente","overdue":"En retard"}
        FEE_L = {"scolarite":"Scolarité","transport":"Transport","cantine":"Cantine","activite":"Activité"}
        total_a = total_p = total_r = 0.0
        for i, fee in enumerate(fees_data, 2):
            amount = float(fee.get("total_amount",0))
            paid   = float(fee.get("paid_amount",0))
            remaining = float(fee.get("remaining_amount", amount - paid))
            total_a += amount; total_p += paid; total_r += remaining
            status = fee.get("status","pending")
            row = [fee.get("student_number",""), fee.get("student_name",""), fee.get("class_name",""),
                   FEE_L.get(fee.get("fee_type",""), fee.get("fee_type","")),
                   amount, paid, remaining, STATUS_L.get(status, status),
                   fee.get("due_date",""), fee.get("paid_date","")]
            for col, val in enumerate(row, 1): ws.cell(row=i, column=col, value=val)
            _data_row(ws, i, len(headers), i % 2 == 0)
            for col in [5,6,7]: ws.cell(row=i, column=col).number_format = '#,##0.00 "MAD"'
            bg, fg = STATUS_C.get(status, (GRAY_LIGHT,"000000"))
            sc = ws.cell(row=i, column=8)
            sc.fill = PatternFill("solid", fgColor=bg)
            sc.font = Font(name="Calibri", bold=True, size=10, color=fg)
            sc.alignment = _center()
            rc = ws.cell(row=i, column=7)
            if remaining > 0:
                rc.fill = PatternFill("solid", fgColor="FFF1F2")
                rc.font = Font(name="Calibri", bold=True, size=10, color=ROSE_HEX)
        tr = len(fees_data) + 2
        for col, val in enumerate(["TOTAL","","","",total_a,total_p,total_r,"","",""], 1):
            cell = ws.cell(row=tr, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=NAVY_HEX)
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.border = _border(); cell.alignment = _center()
        for col in [5,6,7]: ws.cell(row=tr, column=col).number_format = '#,##0.00 "MAD"'
        rate = (total_p/total_a*100) if total_a else 0
        ws.cell(row=tr+1, column=1, value=f"Taux de recouvrement : {rate:.1f}%").font = Font(
            name="Calibri", bold=True, size=10, color=EMERALD_HEX if rate>=80 else ROSE_HEX)
        _title_row(ws, f"Rapport financier — {month or datetime.now().strftime('%B %Y')}", len(headers))
        _freeze_filter(ws, 2, len(headers))
        return _save(wb, output_path)

    def export_grades_class(self, class_data, semester, output_path=None):
        wb = Workbook(); ws = wb.active; ws.title = f"Notes T{semester}"
        students = class_data.get("students",[]); subjects = class_data.get("subjects",[])
        class_name = class_data.get("class_name","Classe")
        headers = ["N° Élève","Nom","Prénom"]
        widths = [13,18,16]
        for subj in subjects:
            headers.append(f"{subj.get('name','')[:12]}\n(coeff {subj.get('coefficient',1)})")
            widths.append(14)
        headers += ["Moy. Gén.","Rang","Mention"]; widths += [12,8,14]
        _header_row(ws, headers, widths)
        def get_mention(avg):
            if avg>=16: return "Très Bien"
            elif avg>=14: return "Bien"
            elif avg>=12: return "Assez Bien"
            elif avg>=10: return "Passable"
            else: return "Insuffisant"
        for i, student in enumerate(students, 2):
            row = [student.get("student_number",""), student.get("last_name","").upper(), student.get("first_name","")]
            for subj in subjects:
                avg = student.get("subject_averages",{}).get(str(subj.get("id","")), None)
                row.append(avg if avg is not None else "—")
            gen_avg = student.get("general_average", 0.0)
            row += [gen_avg, student.get("rank","—"), get_mention(gen_avg)]
            for col, val in enumerate(row, 1): ws.cell(row=i, column=col, value=val)
            _data_row(ws, i, len(headers), i % 2 == 0)
            ac = ws.cell(row=i, column=len(headers)-2)
            ac.number_format = "0.00"; ac.alignment = _center()
            if isinstance(gen_avg,(int,float)):
                ac.fill = PatternFill("solid", fgColor="ECFDF5" if gen_avg>=10 else "FFF1F2")
                ac.font = Font(name="Calibri", bold=True, size=10, color=EMERALD_HEX if gen_avg>=10 else ROSE_HEX)
            for ci in range(4, len(subjects)+4):
                cell = ws.cell(row=i, column=ci)
                cell.alignment = _center(); cell.number_format = "0.00"
                if isinstance(cell.value,(int,float)) and cell.value < 10:
                    cell.fill = PatternFill("solid", fgColor="FFF1F2")
                    cell.font = Font(name="Calibri", size=10, color=ROSE_HEX)
        _title_row(ws, f"Notes — {class_name} — Trimestre {semester}", len(headers))
        _freeze_filter(ws, 2, len(headers))
        return _save(wb, output_path)
